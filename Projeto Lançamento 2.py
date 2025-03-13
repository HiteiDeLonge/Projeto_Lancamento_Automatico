# Este arquivo está sujeito aos termos e condições da Licença de Software Proprietário
# incluída no arquivo LICENSE.md que acompanha este software.

import pyautogui
import sys
import keyboard
from time import sleep
from tkinter import *
import tkinter as tk
from tkinter import Tk, Label, Button, ttk
import xml.etree.ElementTree as ET
from tkinter import filedialog
import openpyxl
from datetime import datetime
from dateutil import parser
import pandas as pd
import os
import re
from reportlab.pdfgen import canvas
from PyPDF2 import PdfReader,PdfWriter
from reportlab.lib.pagesizes import letter
from tkinter import messagebox
from itertools import islice

def popup_menu(event):
    # Verifica se algum item da Treeview foi clicado
    item_id = tree.identify_row(event.y)
    if item_id:
        menu.post(event.x_root, event.y_root)

def limpar_tabela():
    for item in tree.get_children():
        tree.delete(item)
        
def on_vertical_scroll(*args):
    tree.yview(*args)

def on_horizontal_scroll(*args):
    tree.xview(*args)

file_paths = []

def carregar_arquivos():
    # Limpar a árvore antes de adicionar novos dados
    tree.delete(*tree.get_children())

    # Solicitar ao usuário para selecionar arquivos XML
    file_paths = filedialog.askopenfilenames(filetypes=[("Arquivos XML", "*.xml")], title="Selecionar Arquivos XML", multiple=True)

    # Verificar se pelo menos um arquivo foi selecionado
    if not file_paths:
        return

    # Lista para armazenar os arquivos que devem ser excluídos
    arquivos_para_excluir = []

    # Lista para armazenar os dados a serem exibidos na árvore
    dados_para_inserir = []

    # Processar cada arquivo XML
    for i, file_path in enumerate(file_paths):
        try:
            tree_data = processar_xml(file_path)
            
            # Verificar se algum resultado possui "Nome Fantasia não encontrado na planilha"
            if any(resultado[9] == "Nome Fantasia não encontrado na planilha" for resultado in tree_data):
                arquivos_para_excluir.append(file_path)
            else:
                # Adicionar os dados que não precisam ser excluídos
                dados_para_inserir.extend(tree_data)

        except FileNotFoundError:
            print(f"Arquivo não encontrado: {file_path}")
        except Exception as e:
            print(f"Erro ao processar {file_path}: {e}")

    # Excluir os arquivos que possuem "Nome Fantasia não encontrado na planilha"
    for arquivo in arquivos_para_excluir:
        try:
            os.remove(arquivo)
        except Exception as e:
            print(f"Erro ao excluir o arquivo {arquivo}: {e}")

    # Inserir os dados na árvore
    for item in dados_para_inserir:
        var = tk.IntVar()
        tree.insert("", tk.END, values=item)

    # Salvar automaticamente o resultado em um arquivo Excel
    salvar_em_xlsx()
    tree.update()
    janela.update()

def salvar_em_xlsx():
    tree_data = [tree.item(item)["values"] for item in tree.get_children()]

    if not tree_data:
        print("Nenhum dado para salvar.")
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ["numero", "Série", "Emitente CNPJ", "Emissão", "Descrição", "Qnt", "Reais/L", "Total", "Placa","Emitente", "Filial"]
    sheet.append(headers)

    for data in tree_data:
        sheet.append(data)

    # Definir o caminho desejado para salvar o arquivo Excel
    diretorio = r"C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas"
    file_path = os.path.join(diretorio, "notas.xlsx")
    
    # Verificar se o diretório existe, se não, criar o diretório
    if not os.path.exists(diretorio):
        os.makedirs(diretorio)

    workbook.save(file_path)

def formatar_cnpj(cnpj):
    cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    return cnpj_formatado

planilha_de_placas = pd.read_excel(r"C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\planilhas\placas.xlsx")['Placa'].tolist()

planilha_de_nome = pd.read_excel(r"C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas\rs to nm.xlsx")

def processar_xml(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    ns = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

    resultados = {}
    nome_fantasia = "Nome Fantasia não encontrado na planilha"  # Definir o valor padrão

    for prod in root.findall(".//nfe:det", namespaces=ns):
        descricao = prod.find(".//nfe:xProd", namespaces=ns).text
        qnt = float(prod.find(".//nfe:qCom", namespaces=ns).text.replace(',', '.'))
        r_l = float(prod.find(".//nfe:vUnCom", namespaces=ns).text.replace(',', '.'))
        total = float(prod.find(".//nfe:vProd", namespaces=ns).text.replace(',', '.'))

        if "ARLA" in descricao:
            descricao = "ARLA"
        
        if "gasolina" in descricao or "GASOLINA" in descricao:
            descricao = "gasolina"

        if "S10" in descricao or "S-10" in descricao or "S DEZ" in descricao or "S 10" in descricao or "S 1 0" in descricao:
            descricao = "S10"

        if "S500" in descricao or "OLEO DIESEL COMUM" in descricao or "S 500" in descricao or "S-500" in descricao:
            descricao = "S500"

        observacoes = root.find(".//nfe:infAdic/nfe:infCpl", namespaces=ns).text
        placa_match = re.search(r'[A-Z]{3}-?\d[A-Z0-9]-?\d{2}', observacoes)

        # Modificar a linha para obter o nome fantasia do emitente da nota fiscal
        cnpj_emitente = formatar_cnpj(root.find(".//nfe:emit/nfe:CNPJ", namespaces=ns).text)
        if cnpj_emitente in planilha_de_nome['CNPJ'].values:
            # Encontrar o índice onde o CNPJ corresponde
            index = planilha_de_nome['CNPJ'][planilha_de_nome['CNPJ'] == cnpj_emitente].index[0]
            # Obter o nome correspondente na mesma linha
            nome_fantasia = planilha_de_nome.loc[index, 'nome']
        else:
            nome_fantasia = "Nome Fantasia não encontrado na planilha"

        if placa_match:
            placa_encontrada = placa_match.group().upper().replace("-", "")
            if placa_encontrada in planilha_de_placas:
                placa_resultado = placa_encontrada
            else:
                placa_resultado = "Placa não encontrada na planilha"
        else:
            placa_resultado = "Placa não encontrada no campo de observações"

        if descricao in resultados:
            resultados[descricao]['qnt'] += qnt
            resultados[descricao]['r_l'] += r_l
            resultados[descricao]['total'] += total
        else:
            resultados[descricao] = {'qnt': qnt, 'r_l': r_l, 'total': total, 'placa': placa_resultado}

    lista_resultados = [
        (
            root.find(".//nfe:ide/nfe:nNF", namespaces=ns).text,
            root.find(".//nfe:ide/nfe:serie", namespaces=ns).text,
            formatar_cnpj(root.find(".//nfe:emit/nfe:CNPJ", namespaces=ns).text),
            parser.isoparse(root.find(".//nfe:ide/nfe:dhEmi", namespaces=ns).text).strftime("%d/%m/%Y"),
            descricao,
            f'{resultados[descricao]["qnt"]:.2f}',
            f'{resultados[descricao]["r_l"]:.2f}'.replace('.', ','),
            f'{resultados[descricao]["total"]:.2f}'.replace('.', ','),
            placa_resultado,  # Usar a variável placa_resultado ao invés de resultados[descricao]['placa']
            nome_fantasia,  # Usar a variável nome_fantasia definida acima
            formatar_cnpj(root.find(".//nfe:infNFe/nfe:dest/nfe:CNPJ", namespaces=ns).text),
        ) for descricao in resultados
    ]

    return lista_resultados

def rateio():
    # Lendo o arquivo Excel
    excel_path = r'C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas\notas.xlsx'
    df = pd.read_excel(excel_path)

    # Flag para imprimir o cabeçalho apenas uma vez
    print_header = True

    for index, row in df.iterrows():
        nfe = row['numero']
        serie = row['Série']
        modelo = "55"
        cnpj = row['Emitente CNPJ']
        data = str(row['Emissão']).replace('/','')
        operacao = "5015"

        pyautogui.click(1492, 89, duration=0.2)
        pyautogui.click(290, 215, duration=0.2)
        pyautogui.write(str(nfe))
        pyautogui.click(397, 217, duration=0.2)
        pyautogui.write(str(serie))
        pyautogui.click(447, 218, duration=0.2)
        pyautogui.write(modelo)
        pyautogui.press('enter')
        pyautogui.click(438, 88, duration=0.2)
        pyautogui.click(749, 480, duration=0.2)
        pyautogui.write(str(cnpj))
        pyautogui.click(862, 431, duration=0.2)
        pyautogui.doubleClick(334, 258)
        pyautogui.write(data)
        pyautogui.doubleClick(430, 256)
        pyautogui.write(data)
        pyautogui.click(569, 257, duration=0.2)
        pyautogui.write(operacao)
        pyautogui.press('enter')
        pyautogui.click(1561, 90, duration=0.2)
        sleep(3)
        pyautogui.click(1485, 289, duration=0.2)
        sleep(2)

        if print_header:
            print_header = False  
        for index, row in df.iterrows():
            pyautogui.press('enter')
            item = "*" + str(row['Descrição'])
            print(item)
            qnt = str(row['Qnt'])
            print(qnt)
            total = str(row['Total'])
            print(total)
            placa = "*rky6h45"
            print(f"Item: {item}")

            if "ARLA" in item:
                item = "*arla"
                pyautogui.write(item)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(qnt)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(total)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(placa)
                pyautogui.press('enter')         
                pyautogui.click(1556,284,duration=0.2)  
                sleep(2)              
        
            elif "S10" in item:
                item = "*s10"
                pyautogui.write(item)
                pyautogui.press('enter')   
                pyautogui.press('enter')
                pyautogui.write(qnt)
                pyautogui.press('enter')
                pyautogui.press('enter')       
                pyautogui.write(total)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(placa)
                pyautogui.press('enter')         
                pyautogui.click(1556,284,duration=0.2)    
                sleep(2)

            elif "S500" in item or "DIESEL COMUM" in item:
                item = "*s500"
                pyautogui.write(item)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(qnt)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(total)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(placa)
                pyautogui.press('enter')
                pyautogui.click(1556,284,duration=0.2)
                sleep(2)
        break
    pyautogui.press('esc')

def s10():
    pyautogui.click(1068,508)
    pyautogui.click(769,499)
    excel_path = r'C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas\abastecimentos.xlsx'
    df = pd.read_excel(excel_path)

    # Flag para imprimir o cabeçalho apenas uma vez
    print_header = True

    for index, row  in df.iterrows():
        placa_s10 = "*" + (row['Placa'])  
        porcentagem = "{:.3f}".format(row['Diesel S10'])
        print(placa_s10)
        print(porcentagem)
        if porcentagem == '0.000':
            continue
        if porcentagem == '100.0':
            if porcentagem == '0.000':
                continue
            pyautogui.click(1044, 511)
            pyautogui.click(893, 580)
            pyautogui.write(placa_s10)
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.doubleClick(829, 662)
            pyautogui.click(1117, 510)
            break
        pyautogui.click(1044, 511)
        pyautogui.click(893,580)
        pyautogui.write(placa_s10)
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.doubleClick(829, 662)
        pyautogui.write(porcentagem)
        pyautogui.click(1117, 510)
            
        if keyboard.is_pressed('q'):
            break

def s500():
    pyautogui.click(1068,508)
    pyautogui.click(769,499)
    excel_path = r'C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas\abastecimentos.xlsx'
    df = pd.read_excel(excel_path)

    # Flag para imprimir o cabeçalho apenas uma vez
    print_header = True

    for index, row  in df.iterrows():
        placa_s500 = "*" + (row['Placa'])  
        porcentagem ="{:.3f}".format(row['Diesel S500'])
        print(placa_s500)
        print (porcentagem)
        if porcentagem == '0.000':
            continue
        if porcentagem == '100.0':
            if porcentagem == '0.000':
                continue
            pyautogui.click(1044, 511)
            pyautogui.click(893, 580)
            pyautogui.write(placa_s500)
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.doubleClick(829, 662)
            pyautogui.click(1117, 510)
            break
        pyautogui.click(1044, 511)
        pyautogui.click(893,580)
        pyautogui.write(placa_s500)
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.doubleClick(829, 662)
        pyautogui.write(porcentagem)
        pyautogui.click(1117, 510)
        if keyboard.is_pressed('q'):
            break

def arla():
    pyautogui.click(1068,508)
    pyautogui.click(769,499)
    excel_path = r'C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas\abastecimentos.xlsx'
    df = pd.read_excel(excel_path)

    # Flag para imprimir o cabeçalho apenas uma vez
    print_header = True

    for index, row  in df.iterrows():
        placa_arla = "*" + (row['Placa'])  
        porcentagem = "{:.3f}".format(row['Arla Granel'])
        print(placa_arla)
        print (porcentagem)
        if porcentagem == '0.000':
            continue
        if porcentagem == '100.0':
            if porcentagem == '0.000':
                continue
            pyautogui.click(1044, 511)
            pyautogui.click(893, 580)
            pyautogui.write(placa_arla)
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.doubleClick(829, 662)
            pyautogui.click(1117, 510)
            break
        pyautogui.click(1044, 511)
        pyautogui.click(893,580)
        pyautogui.write(placa_arla)
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.doubleClick(829, 662)
        pyautogui.write(porcentagem)
        pyautogui.click(1117, 510)

def imprimir_cabecalho_nota(row):
    nfe = str(row['numero'])
    serie = str(row['Série'])
    modelo = "55"
    cnpj = str(row['Emitente CNPJ'])
    data = str(row['Emissão']).replace('/', '')
    operacao = "5002"

    pyautogui.click(1492, 89, duration=0.2)
    pyautogui.click(290, 215, duration=0.2)
    pyautogui.write(nfe)
    pyautogui.click(397, 217, duration=0.2)
    pyautogui.write(serie)
    pyautogui.click(447, 218, duration=0.2)
    pyautogui.write(modelo)
    pyautogui.press('enter')
    pyautogui.click(438, 88, duration=0.2)
    pyautogui.click(749, 480, duration=0.2)
    pyautogui.write(cnpj)
    pyautogui.click(862, 431, duration=0.2)
    pyautogui.doubleClick(334, 258)
    pyautogui.write(data)
    pyautogui.doubleClick(430, 256)
    pyautogui.write(data)
    pyautogui.click(569, 257, duration=0.2)
    pyautogui.write(operacao)
    pyautogui.press('enter')
    pyautogui.click(1561, 90, duration=0.2)
    sleep(3)
    pyautogui.click(1485, 289, duration=0.2)
    sleep(2)

def imprimir_item_nota(row):
    pyautogui.press('enter')
    descricao_item = str(row['Descrição'])
    qnt = str(row['Qnt'])
    total = str(row['Total'])
    placa = "*" + str(row['Placa'])

    if "ARLA" in descricao_item:
        descricao_item = "*arla"
    elif "S10" in descricao_item:
        descricao_item = "*s10"
    elif "S500" in descricao_item:
        descricao_item = "*s500"
    elif "gasolina" in descricao_item:
        descricao_item = "*gasolina"

    pyautogui.write(descricao_item)
    pyautogui.press('enter')

    if "*arla" in descricao_item:
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.press('enter')

    if "*gasolina" in descricao_item:
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.press('enter')

    if "*s10" in descricao_item or "*s500" in descricao_item:
        pyautogui.press('enter')
    pyautogui.write(qnt)
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.write(total)
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.write(placa)

    if placa == "*CCU7C57" or "*BRY4I72":
        pyautogui.press('enter')

    pyautogui.press('enter')
    pyautogui.click(1556, 284, duration=0.2)
    sleep(3)

def nota_placa():
    caminho_excel = r'C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas\notas.xlsx'
    df = pd.read_excel(caminho_excel)

    # Agrupar por informação da nota fiscal
    df_agrupado = df.groupby(['numero', 'Série', 'Emitente CNPJ', 'Emissão'])

    for (nfe, serie, cnpj, data), grupo_df in df_agrupado:
        # Imprimir cabeçalho para cada nota fiscal
        imprimir_cabecalho_nota(grupo_df.iloc[0])

        # Iterar pelos itens da nota fiscal atual
        for _, row in grupo_df.iterrows():
            imprimir_item_nota(row)

            if keyboard.is_pressed('q'):
                break
        pyautogui.press('esc')

def centralizar_janela(janela):
    largura_tela = janela.winfo_screenwidth()
    altura_tela = janela.winfo_screenheight()

    largura_janela = 800  # Defina a largura da sua janela aqui
    altura_janela = 500   # Defina a altura da sua janela aqui

    x = (largura_tela - largura_janela) // 2
    y = (altura_tela - altura_janela) // 2

    janela.geometry("{}x{}+{}+{}".format(largura_janela, altura_janela, x, y))

def receber():
    while True:
        pyautogui.click(341, 86)
        sleep(3)
        if keyboard.is_pressed('q'):
            break

def on_menu_click():
    # Ação a ser executada quando uma opção do menu é clicada
    selected_item = tree.selection()
    # Adicione aqui o código para a ação desejada

def formatar_numero_nf(numero_nf):
    numero_nf_str = str(numero_nf).zfill(9)
    numero_nf_formatado = '.'.join(numero_nf_str[i:i+3] for i in range(0, 9, 3))
    return numero_nf_formatado

def adicionar_marca_dagua(valores):
    overlay = canvas.Canvas("temp.pdf", pagesize=letter)
    overlay.setFont("Helvetica", 16)
    y = 100
    for prod, icms in valores:
        overlay.drawString(100, y, f"Produto: {prod}, ICMS: {icms}")
        y += 20
    overlay.save()
    print("ICMS CARIMBADO")
    

def encontrar_nota_em_pdf(diretorio_projeto):
    arquivo_excel = os.path.join(diretorio_projeto, r"C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas\icms.xlsx")
    arquivo_pdf = os.path.join(diretorio_projeto, "notas.pdf")

    wb = openpyxl.load_workbook(arquivo_excel)
    planilha = wb.active

    writer = PdfWriter()

    notas_processadas = set()

    with open(arquivo_pdf, 'rb') as pdf_file:
        pdf_reader = PdfReader(pdf_file)

        # Dicionário para armazenar os produtos associados a cada nota fiscal
        produtos_por_nf = {}

        for linha in islice(planilha.iter_rows(values_only=True), 1, None):
            numero_nf, prod, icms = linha[:3]
            numero_nf_formatado = formatar_numero_nf(numero_nf)

            # Adiciona o produto à lista correspondente à nota fiscal
            produtos_por_nf.setdefault(numero_nf_formatado, []).append((prod, icms))

        for numero_nf_formatado, produtos in produtos_por_nf.items():
            # Verifica se a nota já foi processada
            if numero_nf_formatado in notas_processadas:
                continue

            encontrado = False

            for pagina_num, pagina in enumerate(pdf_reader.pages, 1):
                texto = pagina.extract_text()
                if numero_nf_formatado in texto:
                    # Cria uma página com a marca d'água para a nota fiscal
                    adicionar_marca_dagua(produtos)
                    pagina_com_marca = PdfReader("temp.pdf").pages[0]
                    # Adiciona marca d'água para cada produto associado à nota fiscal
                    pagina_com_marca.merge_page(pagina)
                    writer.add_page(pagina_com_marca)
                    encontrado = True
                    break
                
            if not encontrado:
                print(f"NF {numero_nf_formatado} não encontrada no PDF.")

            notas_processadas.add(numero_nf_formatado)

        # Adiciona as páginas originais das notas não processadas
        for pagina in pdf_reader.pages:
            texto = pagina.extract_text()
            if all(nf not in texto for nf in notas_processadas):
                writer.add_page(pagina)

    with open(os.path.join(diretorio_projeto, "notas_com_marca_dagua.pdf"), 'wb') as output_pdf:
        writer.write(output_pdf)

    os.remove("temp.pdf")

diretorio_projeto = r"C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico"

def calcular_icms():
    # Caminho para a planilha de dados
    planilha_de_dados = r"C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas\notas.xlsx"
    
    # Carregar a planilha de dados
    df = pd.read_excel(planilha_de_dados)
    
    # Filtrar os itens que são S10 ou S500
    s10_s500 = df[df['Descrição'].isin(['S10', 'S500'])].copy()  # Copiar o DataFrame para evitar avisos
    
    # Converter a coluna 'Qnt' para numérica
    s10_s500.loc[:, 'Qnt'] = pd.to_numeric(s10_s500['Qnt'], errors='coerce')  # Usar .loc para atribuir valores
    
    # Adicionando uma impressão para verificar a coluna 'Qnt' após a conversão
    print("Valores de Qnt após a conversão:")
    print(s10_s500['Qnt'])
    
    # Calcular o ICMS para os itens selecionados
    s10_s500.loc[:, 'ICMS'] = s10_s500['Qnt'] * 1.0635  # Calcular corretamente o ICMS
    
    # Adicionando uma impressão para verificar os valores de ICMS após o cálculo
    print("\nValores de ICMS após o cálculo:")
    print(s10_s500['ICMS'])
    
    # Selecionar apenas as colunas 'numero', 'Produto' (ou 'Descrição') e 'ICMS'
    s10_s500_icms = s10_s500[['numero', 'Descrição', 'ICMS']]
    
    # Caminho para salvar a planilha de saída
    planilha_de_saida = r"C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas\icms.xlsx"
    
    # Salvar os resultados em uma nova planilha
    s10_s500_icms.to_excel(planilha_de_saida, index=False)
    encontrar_nota_em_pdf(diretorio_projeto)

def criar_tabela_dinamica(arquivo_csv, arquivo_excel):
    # Leitura do arquivo CSV usando ponto e vírgula como separador
    df = pd.read_csv(arquivo_csv, encoding='utf-8', sep=';')

    # Renomeia as colunas para remover espaços e caracteres especiais
    df.columns = df.columns.str.strip().str.replace(' ', '_').str.replace('/', '_')

    # Imprime as colunas do DataFrame
    print("Colunas do DataFrame:")
    print(df.columns)

    try:
        # Substituir ',' por '.' e converter para numérico
        df['Valor_Pago'] = pd.to_numeric(df['Valor_Pago'].str.replace(',', '.'), errors='coerce')
        df['Litros'] = pd.to_numeric(df['Litros'].str.replace(',', '.'), errors='coerce')

        # Criação da tabela dinâmica
        tabela_dinamica = pd.pivot_table(df, values='Valor_Pago', index='Placa', columns='Combustível', aggfunc='sum', fill_value=0, margins=True, margins_name='Total Geral')

        # Calcular porcentagens em relação ao total geral
        tabela_dinamica_porcentagens = tabela_dinamica.div(tabela_dinamica.iloc[-1, :], axis=1) * 100
        tabela_dinamica_porcentagens = tabela_dinamica_porcentagens.iloc[:-1, :]  # Excluir a última linha 'Total Geral'

        # Ajustar o formato das células na tabela de porcentagens
        tabela_dinamica_porcentagens = tabela_dinamica_porcentagens.round(2)

        # Corrigir porcentagens para garantir que a soma seja exatamente 100%
        tabela_dinamica_porcentagens = tabela_dinamica_porcentagens * (100 / tabela_dinamica_porcentagens.sum())

        # Escrever a tabela dinâmica e porcentagens em um arquivo Excel
        with pd.ExcelWriter(arquivo_excel, engine='xlsxwriter') as writer:
            tabela_dinamica_porcentagens.to_excel(writer, sheet_name='Porcentagens', index=True)

            # Obter as folhas do Excel
            workbook = writer.book
            worksheet_p = writer.sheets['Porcentagens']

            # Formato padrão para as células
            formato_padrao = workbook.add_format({'bold': False, 'num_format': '#,##0.00'})

            # Formato para porcentagens com duas casas decimais
            formato_porcentagem = workbook.add_format({'num_format': '0.00%'})

            # Aplicar formato padrão a todas as células em ambas as folhas
            for worksheet in [worksheet_p]:
                for col_num, value in enumerate(tabela_dinamica.columns.get_level_values('Combustível')):
                    worksheet.set_column(col_num, col_num, None, formato_padrao)

                if 'Porcentagens' in worksheet.name:
                    for col_num, value in enumerate(tabela_dinamica_porcentagens.columns):
                        worksheet.set_column(col_num + len(tabela_dinamica.columns), col_num + len(tabela_dinamica.columns), None, formato_porcentagem)

        # Imprime a tabela dinâmica e porcentagens
        print(tabela_dinamica)
        
        print("\nTabela com Porcentagens:")
        print(tabela_dinamica_porcentagens)
        
        print(f"\nTabelas criadas e salvas em {arquivo_excel}")
    except Exception as e:
        messagebox.showinfo("Erro ao criar as porcentagens",f"Erro ao criar a tabela dinâmica: {e}")


diretorio_projeto = r"C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico"

def carregar_abastecimentos():
    arquivo_csv = filedialog.askopenfilename(filetypes=[("Arquivos CSV", "*.csv")])

    if not arquivo_csv:
        print("Por favor, selecione um arquivo CSV.")
        return
    
    # Definir o caminho desejado para salvar o arquivo Excel
    caminho_salvamento = "C:/Users/Log20-2/Desktop/Projeto Lançamento Automatico/Planilhas/abastecimentos.xlsx"
    
    # Verificar se o diretório existe, se não, criar o diretório
    diretorio = os.path.dirname(caminho_salvamento)
    if not os.path.exists(diretorio):
        os.makedirs(diretorio)
    
    # Chamar a função para criar a tabela dinâmica
    criar_tabela_dinamica(arquivo_csv, caminho_salvamento)

def abrir_planilha_notas():
    caminho_arquivo = r'C:\Users\Log20-2\Desktop\Projeto Lançamento Automatico\Planilhas\notas.xlsx'
    try:
        os.startfile(caminho_arquivo)
        print("Arquivo aberto com sucesso!")
    except FileNotFoundError:
        print("O arquivo não foi encontrado.")

janela = Tk()
janela.title("Sistema de Lançamentos")
janela.geometry("800x500")
janela.configure(background="#ffffff")

#BOTÕES
botao = Button(janela, text="Rateio", command=rateio, background="#ffffff")
botao.config(font=("Arial", 16))
botao.place(x=29, y=209, width=128, height=59)
botao2=Button(janela, text="Placa", command=nota_placa, background="#ffffff")
botao2.config(font=("Arial", 16))
botao2.place(x=29, y=300, width=128, height=59)
botao4=Button(janela, text="Diesel S10", command=s10, background="#ffffff")
botao4.place(x=409, y=74, width=128,height=59)
botao4.config(font=("Arial", 16))
botao5=Button(janela, text="Arla", command=arla, background="#ffffff")
botao5.place(x=246, y=74, width=128,height=59)
botao5.config(font=("Arial", 16))
botao6=Button(janela, text="S500", command=s500, background="#ffffff")
botao6.place(x=572, y=74, width=128,height=59)
botao6.config(font=("Arial", 16))
devolpedby=Label(janela, text="Desenvolvido por Bernardo Vilbert", font=("Arial", 12), background="#ffffff", foreground="#000000")
devolpedby.place(x=280, y=470)
botao7=Button(janela, text="Receber", command=receber, background="#ffffff")
botao7.place(x=29, y=391, width=128, height=59)
botao7.config(font=("Arial, 16"))
botao_carregar = Button(janela, text="Carregar Dados", command=carregar_arquivos, background="#ffffff")
botao_carregar.place(x=10, y=10)
menu = tk.Menu(janela, tearoff=0)
menu.add_command(label="Rateio", command=rateio, background="#ffffff")
menu.add_command(label="Nota Placa", command=nota_placa, background="#ffffff")
menu.add_command(label="Abrir Notas", command=abrir_planilha_notas,background="#ffffff")
botao_limpar = tk.Button(janela, text="Limpar Tabela", command=limpar_tabela, background="#ffffff")
botao_limpar.place(x=10, y=50)
botao_abastecidas = tk.Button(janela, text="Carregar abastecimentos", command=carregar_abastecimentos, background="#ffffff")
botao_abastecidas.place(x=10, y=90)
botao_icms = tk.Button(janela,text="Carimbar ICMS",command=calcular_icms,background="#ffffff")
botao_icms.place(x=10,y=130)


#CRIANDO TABELA
tree = ttk.Treeview(janela, columns=("numero", "Série", "Emitente CNPJ","Emissão","Descrição","Qnt","Reais/L","Total","Placa","Emitente","Filial"))
tree.heading("numero", text="Número")
tree.heading("Série", text="Série")
tree.heading("Emitente CNPJ", text="Emitente CNPJ")
tree.heading("Emissão", text="Emissão")
tree.heading("Descrição", text = "Descrição")
tree.heading("Qnt", text = "Qnt")
tree.heading("Reais/L", text = "Reais/L")
tree.heading("Total", text = "Total")
tree.heading("Placa", text = "Placa")
tree.heading("Emitente", text = "Emitente")
tree.heading("Filial", text="Filial")
tree.column("numero", width=80)
tree.column("Série", width=45)
tree.column("Emitente CNPJ", width=120)
tree.column("Emissão", width=120)
tree.column("Descrição", width=80)
tree.column("Qnt", width=80)
tree.column("Reais/L", width=50)
tree.column("Total", width=80)
tree.column("Placa", width=80)
tree.column("Emitente", width=150)
tree.column("Filial", width=150)
tree.place(x=199, y=141, width=531, height=285)
tree.bind("<Button-3>", popup_menu)
tree["show"] = "headings"

#BARRAS DE ROLAGEM DA TABELA
vertical_scrollbar = ttk.Scrollbar(janela, orient="vertical", command=on_vertical_scroll)
vertical_scrollbar.place(x=730, y=141, height=285)
horizontal_scrollbar = ttk.Scrollbar(janela, orient="horizontal", command=on_horizontal_scroll)
horizontal_scrollbar.place(x=199, y=426, width=531)
tree.configure(yscrollcommand=vertical_scrollbar.set, xscrollcommand=horizontal_scrollbar.set)


#INICIALIZANDO COM O SISTEMA
janela.update_idletasks()
centralizar_janela(janela)
janela.update()
janela.mainloop()
